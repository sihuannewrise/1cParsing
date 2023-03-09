# from json import load
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

CWD = os.getcwd()

SOURCE_DIR = './files/in/'
DESTIN_DIR = './files/out/'

SOURCE_FILE_NAME = 'knh51-2022.xlsx'

DESTIN_FILE_SUFFIX = ' - out'
DESTIN_SHEET_NAME = 'proc'

file = SOURCE_DIR + SOURCE_FILE_NAME
PERIOD, VAT, DATA, four, five, six, AMOUNT = range(7)


def avd(key, val, dict, is_last):
    if is_last:
        if key in dict:
            dict[key] += val
        else:
            dict.update({key: val})
            print(f'{dict}')
    else:
        if key not in dict:
            dict.update({key: val})
    return dict


def collect_vocabulary(filename):
    """Загружаем файл.
    """
    wb = load_workbook(filename=filename)
    ws = wb.active
    # months = [0 for _ in range(12)]
    # contract = {}
    ca = dict()

    for row in ws.iter_rows(min_row=2, max_row=3, values_only=True):
        current_date = datetime.strptime(row[PERIOD], '%d.%m.%Y').date()
        year, current_month = current_date.strftime('%Y'), current_date.strftime('%m')
        _, current_ca, current_contract, *_ = row[DATA].split('\n')
        current_sum = round(float(row[AMOUNT]), 2)

        if current_ca not in ca:
            ca.update({current_ca: {}})
            print(ca)
            if current_contract not in ca[current_ca]:
                ca.update({current_ca: {current_contract: {}}})
                print(ca)
                if current_month not in ca[current_ca][current_contract]:
                    ca.update({current_ca: {current_contract: {current_month: current_sum}}})
                else:
                    ca[current_ca][current_contract][current_month] += current_sum

        # ca = avd(
        #     current_ca,
        #     avd(
        #         current_contract,
        #         avd(
        #             current_month, current_sum, dict(), True),
        #         dict(), False),
        #     dict(), False)
    return ca


print(collect_vocabulary(file))

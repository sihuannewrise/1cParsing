import os
import re
import json

from openpyxl import load_workbook, Workbook

CWD = os.getcwd()

vat_search = r'Без н'
ip_search = r'ИП'
zero_vat_ca = set()

SOURCE_DIR = 'app/files/in/'
DESTIN_DIR = 'app/files/'

SOURCE_FILE_NAME = 'knh51-2022.xlsx'
DESTIN_FILE_NAME = 'expenses.xlsx'

file = SOURCE_DIR + SOURCE_FILE_NAME
xlsx_file = DESTIN_DIR + DESTIN_FILE_NAME
HEADER = ['№ п/п', 'Контрагент', 'Группа', 'Договор', 'Сумма за год']
PERIOD, VAT, DATA, four, five, six, AMOUNT = range(7)


class Expenses(dict):
    def __missing__(self, key):
        value = self[key] = type(self)()
        return value


def create_dict(filename) -> dict:
    """Загружаем файл.
    """
    wb = load_workbook(filename=filename)
    ws = wb.active
    expenses = Expenses()

    for row in ws.iter_rows(min_row=2, values_only=True):
        # current_date = datetime.strptime(row[PERIOD], '%d.%m.%Y').date()
        # year = current_date.strftime('%Y')
        # current_month = int(current_date.strftime('%m'))
        _, current_ca, current_contract, *_ = row[DATA].split('\n')
        zero_vat = re.search(vat_search, row[VAT])
        ip_ca = re.search(ip_search, current_ca)
        if zero_vat or ip_ca:
            current_sum = int(row[AMOUNT])
            zero_vat_ca.add(current_ca)
        else:
            current_sum = int(row[AMOUNT]/1.2)

        # с разбивкой по годам
        if current_contract not in expenses[current_ca]:
            expenses[current_ca][current_contract] = current_sum
        else:
            expenses[current_ca][current_contract] += current_sum

        # с разбивкой по месяцам
        # if current_month not in expenses[current_ca][current_contract]:
        #     expenses[current_ca][current_contract][current_month] = round(
        #         current_sum/1.2, 2)
        # else:
        #     expenses[current_ca][current_contract][current_month] += round(
        #         current_sum/1.2, 2)
    print(zero_vat_ca)
    return expenses


# exp = create_dict(file)
# print(type(exp), len(exp), exp)


def write_dict_to_json(data):
    with open("app/files/expenses.json", "w", encoding='UTF-8') as outfile:
        json.dump(data, outfile, ensure_ascii=False)
    return None


# write_dict_to_json(create_dict(file))


def write_dict_to_xlsx(data):
    wb = Workbook()
    ws = wb.active

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['D'].width = 44
    ws.column_dimensions['E'].width = 12
    ws.freeze_panes = 'B2'
    ws.append(HEADER)

    c_col, c_row = 1, 2
    line_delta = c_row - 1
    for ca in data:
        for contract in data[ca]:
            ws.cell(column=c_col, row=c_row).value = c_row - line_delta
            ws.cell(column=c_col + 1, row=c_row).value = ca
            ws.cell(column=c_col + 3, row=c_row).value = contract
            ws.cell(column=c_col + 4, row=c_row).value = data[ca][contract]
            c_row += 1
    wb.save(filename=xlsx_file)


write_dict_to_xlsx(create_dict(file))
